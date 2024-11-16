import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Function to process files and thresholds
def process_files(month_data, thresholds):
    results = {}
    fail_summary = pd.DataFrame()

    for month, data in month_data.items():
        if data is None:  # Skip if the file is not uploaded
            continue
        try:
            # Read the required sheet starting from the correct header row
            sheet_data = pd.read_excel(data, sheet_name="Total Hour Calculation", header=2)
            # Extract necessary columns
            site_kpi = sheet_data[["Site ID", "Site wise KPI"]]
            # Handle missing KPI values
            missing_kpi = site_kpi[site_kpi["Site wise KPI"].isna()]
            site_kpi = site_kpi.dropna(subset=["Site wise KPI"])

            # Add Threshold and Pass/Fail columns
            site_kpi["Threshold"] = thresholds[month]
            site_kpi["Pass/Fail"] = site_kpi["Site wise KPI"].apply(
                lambda x: "Pass" if x >= thresholds[month] else "Fail"
            )

            # Add Fail results to the summary
            site_kpi["Month"] = month
            fail_summary = pd.concat(
                [fail_summary, site_kpi[site_kpi["Pass/Fail"] == "Fail"]],
                ignore_index=True
            )

            results[month] = (site_kpi, missing_kpi)
        except KeyError as e:
            st.error(f"Error processing {month}: Missing required columns. {e}")
        except Exception as e:
            st.error(f"Unexpected error with {month}: {e}")

    return results, fail_summary

# Function to analyze fail results
def analyze_fails(fail_summary):
    fail_count = fail_summary.groupby("Site ID")["Pass/Fail"].count()
    fail_consecutive = fail_summary.sort_values(["Site ID", "Month"])
    fail_consecutive["Consecutive Fail"] = (
        fail_consecutive.groupby("Site ID").cumcount() + 1
    )

    # Sites failing 5 or more times
    frequent_fails = fail_count[fail_count >= 5].reset_index()
    frequent_fails.columns = ["Site ID", "Fail Count"]

    # Sites failing for 3 consecutive months
    consecutive_fails = fail_consecutive[
        fail_consecutive["Consecutive Fail"] >= 3
    ].drop_duplicates(subset=["Site ID"])[["Site ID"]]

    return frequent_fails, consecutive_fails

# Streamlit App
st.title("Monthly KPI Comparison Tool with Fail Analysis")

# Initialize month names
months = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

# Step 1: File upload and threshold input
month_data = {}
thresholds = {}

st.sidebar.header("Upload Files and Set Thresholds")
for month in months:
    st.sidebar.subheader(f"{month}")
    uploaded_file = st.sidebar.file_uploader(f"Upload {month} File", type=["xlsx"], key=month)
    if uploaded_file is not None:
        month_data[month] = uploaded_file
        thresholds[month] = st.sidebar.number_input(f"{month} KPI Threshold", min_value=0.0, value=0.0)
    else:
        st.sidebar.write(f"Ignoring {month}.")

# Step 2: Process data when "Process" is clicked
if st.button("Process Files"):
    # Ensure at least one file is uploaded
    if all(data is None for data in month_data.values()):
        st.warning("Please upload at least one file!")
    else:
        results, fail_summary = process_files(month_data, thresholds)
        if results:
            # Analyze fails
            frequent_fails, consecutive_fails = analyze_fails(fail_summary)

            # Combine results into a single Excel workbook
            wb = Workbook()
            for month, (processed_data, missing_data) in results.items():
                ws = wb.create_sheet(title=month)
                for row in dataframe_to_rows(processed_data, index=False, header=True):
                    ws.append(row)

                # Highlight missing KPI values in red
                for row in dataframe_to_rows(missing_data, index=False, header=True):
                    ws.append(row)
                    for cell in ws[-1]:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # Add summary sheet
            summary_ws = wb.create_sheet(title="Fail Summary")
            for row in dataframe_to_rows(frequent_fails, index=False, header=True):
                summary_ws.append(row)

            # Save the workbook
            wb.save("KPI_Results_with_Analysis.xlsx")

            # Show tables in Streamlit
            st.subheader("Sites Failing 5 or More Times")
            st.write(frequent_fails)

            st.subheader("Sites Failing for 3 Consecutive Months")
            st.write(consecutive_fails)

            # Provide download button
            with open("KPI_Results_with_Analysis.xlsx", "rb") as f:
                st.download_button("Download Results", data=f, file_name="KPI_Results_with_Analysis.xlsx")
        else:
            st.warning("No files were processed.")
